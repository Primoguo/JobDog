""""
Excel API - Python implementation using openpyxl
基于 openpyxl 实现的 Excel API

API USAGE NOTES:
- When working with cells, rows, and columns, use Excel-style addresses (e.g., "A1", "B5")
  instead of row and column indices.
- You can refer to cells and ranges across sheets by using the sheet name and the cell address,
  e.g. "Sheet1!A1" or "Sheet2!B2:C4".
- Any method that takes an Excel-style range or address can only take a single range or address,
  not multiples (e.g., avoid "A1:B2,C3:D4"). If you need to work with multiple ranges,
  do them one at a time or combine them into a single range if they are contiguous.
"""
import json
import shutil
from openpyxl import Workbook as OpenpyxlWorkbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers, Protection
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string, range_boundaries
from typing import Optional, Dict, Any, List, Tuple, Union
from enum import Enum
import re
from copy import copy

try:
    from .auto_resize import format_cell_value, calculate_display_width
    from .recalc import recalc
except ImportError:
    from auto_resize import format_cell_value, calculate_display_width
    from recalc import recalc

class HorizontalAlign(Enum):
    """水平对齐方式枚举"""
    LEFT = 'left'
    CENTER = 'center'
    RIGHT = 'right'
    GENERAL = 'general'


class VerticalAlign(Enum):
    """垂直对齐方式枚举"""
    TOP = 'top'
    CENTER = 'center'
    BOTTOM = 'bottom'


class LineStyle(Enum):
    """边框线型枚举"""
    THIN = 'thin'
    MEDIUM = 'medium'
    DASHED = 'dashed'
    DOTTED = 'dotted'
    THICK = 'thick'
    DOUBLE = 'double'
    HAIR = 'hair'
    MEDIUM_DASHED = 'mediumDashed'
    DASH_DOT = 'dashDot'
    MEDIUM_DASH_DOT = 'mediumDashDot'
    DASH_DOT_DOT = 'dashDotDot'
    MEDIUM_DASH_DOT_DOT = 'mediumDashDotDot'
    SLANTED_DASH_DOT = 'slantDashDot'

class Worksheet:
    """工作表封装类"""

    def __init__(self, sheet: OpenpyxlWorksheet, workbook: 'Workbook'):
        self._sheet = sheet
        self._workbook = workbook

    def set_name(self, name: str):
        """
        设置工作表名称

        Args:
            name: 新的工作表名称
        """
        self._sheet.title = name

    def get_name(self) -> str:
        """
        获取工作表名称

        Returns:
            工作表名称
        """
        return self._sheet.title

    def get_index(self) -> int:
        """
        获取工作表索引（从0开始）

        Returns:
            工作表在工作簿中的索引位置
        """
        return self._workbook._wb.index(self._sheet)

    def get_row_count(self) -> int:
        """
        获取总行数

        Returns:
            工作表中的总行数，即使这些行不在上下文中
        """
        return self._sheet.max_row

    def get_column_count(self) -> int:
        """
        获取总列数

        Returns:
            工作表中的总列数，即使这些列不在上下文中
        """
        return self._sheet.max_column

    def get_max_row_with_data_in_column(self, column_range: str) -> int:
        """
        获取指定范围（可以包含整列，如"A:E"或"A1:E10"）内的有值单元格的最大行号。

        Args:
            column_range: Excel column范围字符串，例如 "A:E", "A", "A1:E10", "A1"。

        Returns:
            在该范围内的列中找到的有值单元格最大有效行号。如果指定列范围内没有有效数据，则返回0。

        Example:
            >>> # 假设 Sheet1 的 A1:C5 有数据，A:C 列的最后一行数据在 C10
            >>> sheet.get_max_row_with_data_in_column("A:C") # 返回 10 (或更高，如果 max_row 导致)
            >>> sheet.get_max_row_with_data_in_column("A") # 返回 A 列找到的最高有效行
        """
        max_row_in_range = 0
        min_col, _, max_col, _ = range_boundaries(column_range)
        for c_idx in range(min_col, max_col + 1):
            for r_idx in range(self._sheet.max_row, 0, -1):
                cell = self._sheet._cells.get((r_idx, c_idx))
                # 检查单元格是否存在且有值
                if cell and cell.value is not None:
                    max_row_in_range = max(max_row_in_range, r_idx)
                    break
        return max_row_in_range

    def add_rows_at(self, row_address: str, count: int):
        """
        在指定行插入行

        Args:
            row_address: 行地址，如 "5" 表示在第5行插入
            count: 要插入的行数
        """
        row = int(re.findall(r'\d+', row_address)[0])
        self._sheet.insert_rows(row, count)

    def delete_rows_at(self, row_address: str, count: int):
        """
        删除指定行

        Args:
            row_address: 行地址，如 "5" 表示删除第5行
            count: 要删除的行数
        """
        row = int(re.findall(r'\d+', row_address)[0])
        self._sheet.delete_rows(row, count)

    def add_columns_at(self, column_letter: str, count: int):
        """
        在指定列插入列

        Args:
            column_letter: 列字母，如 "C"
            count: 要插入的列数
        """
        col = column_index_from_string(column_letter)
        self._sheet.insert_cols(col, count)

    def delete_columns_at(self, column_letter: str, count: int):
        """
        删除指定列

        Args:
            column_letter: 列字母，如 "C"
            count: 要删除的列数
        """
        col = column_index_from_string(column_letter)
        self._sheet.delete_cols(col, count)

    def shift_up_cells(self, range_str: str, max_row_with_data_in_column: Optional[int] = None) -> str:
        """
        实现 Excel 中的 "删除并上移单元格" 功能。
        删除指定范围内的单元格内容和格式，并将下方单元格的内容和格式向上移动。
        同时处理合并单元格，保留其格式和设置。

        **合并单元格处理规则:**
        1.  **和删除范围有重叠的合并单元格：**
            *   如果删除范围 `range_str` 未完全包裹一个合并单元格（即合并单元格在行或列上溢出 `range_str`），将抛出错误。
            *   如果合并单元格完全包含在 `range_str` 内，它将被删除且不进行特殊处理。
        2.  **完全在删除范围下方的合并单元格：**
            *   如果删除范围下方存在一个合并单元格，且其列宽与 `range_str` 的列宽完全重叠，或完全包含在 `range_str` 内，则该合并单元格及其内容将整体向上移动相应的行数。
            *   如果删除范围下方存在一个合并单元格，但其列宽与 `range_str` 的列宽仅部分重叠，或不能完全包含在`range_str` 内，将抛出错误。
        3.  **完全在删除范围上方的合并单元格：**
            *   删除范围上方的合并单元格不受任何影响

        Args:
            range_str: 要删除并上移的单元格范围，例如 "A1:B5"，或单个单元格如 "A1"
            max_row_with_data_in_column: 指定要上移区域所属列的最后一行，注意不是整个sheet的max_row。如果不传入，则默认为 `self.get_max_row_with_data_in_column(range_str)`。传入该参数可以限制内部搜索和操作的范围，避免执行超时。

        Returns:
            操作结果的字符串描述，例如 "成功将区域 A1:B5 的内容上移。" 或包含警告信息。

        Raises:
            ValueError: 如果合并单元格配置与上移操作冲突。

        **性能提示 (Performance Note):**
        - 如果需要大量调用 `shift_up_cells` 方法来删除连续的行
            - 必须将这些连续的删除区域合并为一个大的 `range_str` 进行调用，而不是多次调用小的区域。禁止在for循环中不加区域合并逻辑的无脑调用该方法
            - 必须设置 `max_row_with_data_in_column` 参数，可以在调用该方法前使用 `get_max_row_with_data_in_column()` 方法获取这些列的最大有效行数。
        - 例如，要删除 "B2:C2" 和 "B3:C3"，应该合并为 `shift_up_cells("B2:C3")` 一次调用，而不是 `shift_up_cells("B3:C3")` 后紧接着 `shift_up_cells("B2:C2")`。

        Example:
            >>> # 删除 A1:B2，将 A3:B_max 区域的内容上移到 A1:B_max-2
            >>> sheet.shift_up_cells("A1:B2", 100) # 知道实际有效数据不会超过 100 行
        """
        min_col, min_row, max_col, max_row = General._parse_range(range_str)
        shift_height = max_row - min_row + 1  # 需要上移的行数

        if shift_height <= 0:
            return f"指定的范围 '{range_str}' 无效或为空，未执行任何上移操作。"

        current_merged_ranges = list(self._sheet.merged_cells.ranges)
        # print(current_merged_ranges)
        merges_to_unmerge_and_remerge = []  # 需要上移的合并单元格
        ranges_to_unmerge_in_place = []  # 需要在此次操作中取消合并的（例如，完全被删除的）
        warnings = []

        # Helper for intersection
        def _intersects(r1_min_c, r1_min_r, r1_max_c, r1_max_r, r2_min_c, r2_min_r, r2_max_c, r2_max_r) -> bool:
            return not (r1_max_r < r2_min_r or r1_min_r > r2_max_r or
                        r1_max_c < r2_min_c or r1_min_c > r2_max_c)

        # Helper for containment
        def _is_contained(inner_min_c, inner_min_r, inner_max_c, inner_max_r, outer_min_c, outer_min_r, outer_max_c,
                          outer_max_r) -> bool:
            return (inner_min_c >= outer_min_c and inner_max_c <= outer_max_c and
                    inner_min_r >= outer_min_r and inner_max_r <= outer_max_r)

        # 1. 预检查所有合并单元格，进行冲突判断并识别需要移动的合并单元格
        for m_range_obj in current_merged_ranges:
            m_min_col, m_min_row, m_max_col, m_max_row = General._parse_range(m_range_obj.coord)

            # 获取删除范围的坐标
            delete_range_coords = (min_col, min_row, max_col, max_row)
            # 获取当前合并单元格的坐标
            merged_cell_coords = (m_min_col, m_min_row, m_max_col, m_max_row)

            # C1: 合并单元格与删除范围部分重叠，且不被删除范围完全包含 -> 报错
            # 这涵盖了 "有合并单元格未完全包裹进去" 的情况
            if _intersects(*delete_range_coords, *merged_cell_coords) and \
                    not _is_contained(*merged_cell_coords, *delete_range_coords):
                raise ValueError(
                    f"无法执行 '删除并上移单元格' 操作。合并单元格 '{m_range_obj.coord}' "
                    f"与删除范围 '{range_str}' 部分重叠，不能被完全包含或完全排除。"
                )

            # C2: 合并单元格完全在删除范围之下 (行不重叠, m_min_row > max_row)
            if m_min_row > max_row:
                # 检查列是否完全重叠
                if m_min_col <= min_col and m_max_col <= max_col:  # 合并单元格的列完全等于删除区域的列，或完全包含在删除区域内
                    # 列完全重叠或完全包含，此合并单元格需要上移
                    merges_to_unmerge_and_remerge.append(m_range_obj)
                else:
                    # 列部分重叠，报错
                    raise ValueError(
                        f"无法执行 '删除并上移单元格' 操作。合并单元格 '{m_range_obj.coord}' "
                        f"在删除范围 '{range_str}' 之下，但列不完全重叠或未包含在删除区域的列内。 "
                        f"(删除区域列: {get_column_letter(min_col)}-{get_column_letter(max_col)} vs "
                        f"合并单元格列: {get_column_letter(m_min_col)}-{get_column_letter(m_max_col)})。"
                    )
            # C3: 合并单元格完全在删除范围内 (将被覆盖/清除，不需要上移)
            elif _is_contained(*merged_cell_coords, *delete_range_coords):
                ranges_to_unmerge_in_place.append(m_range_obj)
            # C4: 合并单元格完全在删除范围之上 (行不重叠, m_max_row < min_row)
            # 这种情况下，合并单元格不受影响，无需操作。

        # 2. 移除所有将被移动或被清除的合并单元格
        # openpyxl 的 unmerge_cells 接受字符串坐标
        for m_range_obj in merges_to_unmerge_and_remerge + ranges_to_unmerge_in_place:
            try:
                self._sheet.unmerge_cells(m_range_obj.coord)
            except KeyError:
                # 理论上不会发生，但在复杂场景下可能出现重复unmerge等情况，做个简单防护
                pass

        # 3. 执行单元格内容和格式的上移
        max_current_row = self._sheet.max_row
        if max_row_with_data_in_column is not None and max_row_with_data_in_column < max_current_row: # 如果传入了比max_row小的值，则使用传入的值
            max_current_row = max_row_with_data_in_column
        else:
            max_current_row = self.get_max_row_with_data_in_column(range_str) # 否则获取该区域列的最大值

        # 从目标区域的第一行开始，到能够作为源区域的最后一行
        # 范围是 [min_row, max_current_row - shift_height]
        for r in range(min_row, max_current_row - shift_height + 1):
            for c in range(min_col, max_col + 1):
                source_cell = self._sheet.cell(row=r + shift_height, column=c)
                target_cell = self._sheet.cell(row=r, column=c)
                target_cell.value = source_cell.value
                # 复制样式
                if source_cell.has_style:
                    # 必须使用copy()来防止修改原始样式对象
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.protection = copy(source_cell.protection)
                else:
                    # 如果源单元格没有样式，则清除目标单元格的样式
                    target_cell.font = Font()
                    target_cell.border = Border()
                    target_cell.fill = PatternFill()
                    target_cell.number_format = numbers.FORMAT_GENERAL
                    target_cell.alignment = Alignment()
                    target_cell.protection = Protection()

        # 4. 清除底部多余的单元格
        # 这些是原来最底部的 shift_height 行，其中现在保存的是重复数据
        # 循环 r 从 max_current_row - shift_height + 1 到 max_current_row
        for r in range(max_current_row - shift_height + 1, max_current_row + 1):
            for c in range(min_col, max_col + 1):
                cell = self._sheet.cell(row=r, column=c)
                cell.value = None
                # 清除格式
                cell.font = Font()
                cell.border = Border()
                cell.fill = PatternFill()
                cell.number_format = numbers.FORMAT_GENERAL
                cell.alignment = Alignment()
                cell.protection = Protection()

        # 5. 重新合并之前标记为需要上移的合并单元格
        remerged_count = 0
        for m_range_obj in merges_to_unmerge_and_remerge:
            new_min_row = m_range_obj.min_row - shift_height
            new_max_row = m_range_obj.max_row - shift_height
            # 确保新的合并范围仍然有效（即起始行不再小于1，且结束行不小于起始行）
            if 1 <= new_min_row <= new_max_row:
                new_coord = f"{get_column_letter(m_range_obj.min_col)}{new_min_row}:{get_column_letter(m_range_obj.max_col)}{new_max_row}"
                self._sheet.merge_cells(new_coord)
                remerged_count += 1
            else:
                warnings.append(f"合并区域 '{m_range_obj.coord}' 向上移动后超出工作表边界或变为无效，已跳过重新合并。")

        result_message = f"成功将区域 '{range_str}' 的内容向上移动 {shift_height} 行。"
        if remerged_count > 0:
            result_message += f" 成功移动并重新合并了 {remerged_count} 个合并单元格。"
        if warnings:
            result_message += " 警告：" + " ".join(warnings)

        return result_message

    def auto_resize_column_at(self, column_letter: str):
        """
        自动调整列宽以适应内容

        Args:
            column_letter: 列字母，如 "A"
        """
        col = column_index_from_string(column_letter)
        max_length = 0
        for row in self._sheet.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    # 格式化显示值
                    cell_text = format_cell_value(cell)
                    # 获取字体属性
                    font_size = float(cell.font.size) if cell.font and cell.font.size else 11.0
                    bold = bool(cell.font.bold) if cell.font else False
                    cell_length = calculate_display_width(cell_text, font_size, bold)
                    max_length = max(max_length, cell_length)
        adjusted_width = max_length + 2  # buffer ≈ 20 像素
        adjusted_width = max(adjusted_width + 2, 8)  # padding
        adjusted_width = min(adjusted_width, 50)  # 限制最大宽度
        adjusted_width = round(adjusted_width, 1)
        self._sheet.column_dimensions[column_letter].width = adjusted_width

    def auto_resize_columns(self):
        """
        自动调整当前工作表所有列的列宽，以适应其内容。
        它会根据单元格内容的长度、字体大小和加粗情况来计算列宽。
        """
        worksheet = self._sheet
        column_widths = {}
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    column_letter = get_column_letter(cell.column)
                    # 格式化显示值
                    cell_text = format_cell_value(cell)
                    # 获取字体属性
                    font_size = float(cell.font.size) if cell.font and cell.font.size else 11.0
                    bold = bool(cell.font.bold) if cell.font else False
                    cell_length = calculate_display_width(cell_text, font_size, bold)
                    if column_letter not in column_widths:
                        column_widths[column_letter] = cell_length
                    else:
                        column_widths[column_letter] = max(
                            column_widths[column_letter],
                            cell_length
                        )

        # 应用列宽
        for column_letter, width in column_widths.items():
            adjusted_width = width + 2  # buffer ≈ 20 像素
            adjusted_width = max(adjusted_width + 2, 8)  # padding
            adjusted_width = min(adjusted_width, 50)  # 限制最大宽度
            adjusted_width = round(adjusted_width, 1)

            worksheet.column_dimensions[column_letter].width = adjusted_width
            print(f"Sheet {self._sheet.title} Column {column_letter}: width set to {adjusted_width}")

    def visible(self, show: bool):
        """
        显示或隐藏整个工作表

        Args:
            show: True 显示工作表，False 隐藏工作表
        """
        self._sheet.sheet_state = 'visible' if show else 'hidden'

    def get_used_range(self) -> str:
        """
        获取已使用的范围

        Returns:
            已使用范围的地址，如 "A1:D10"
        """
        if self._sheet.max_row == 0 or self._sheet.max_column == 0:
            return "A1"
        return f"A1:{get_column_letter(self._sheet.max_column)}{self._sheet.max_row}"

    def regex_search(self, patterns: List[str], match_case: bool = False) -> List[Dict[str, Any]]:
        """
        使用正则表达式搜索单元格

        Args:
            patterns: 正则表达式模式列表
            match_case: 是否区分大小写，默认 False

        Returns:
            匹配结果列表，每个结果包含 'address' 和 'value' 键
        """
        results = []
        flags = 0 if match_case else re.IGNORECASE
        compiled_patterns = [re.compile(p, flags) for p in patterns]

        for row in self._sheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell_str = str(cell.value)
                    for pattern in compiled_patterns:
                        if pattern.search(cell_str):
                            results.append({
                                'address': cell.coordinate,
                                'value': cell.value
                            })
                            break
        return results

    def get_sheet_summary(self, include_data: bool = False) -> str:
        """
        获取工作表概览，包括名称、维度、已使用范围

        Args:
            include_data: 如果为 True，则包含工作表中的数据（前10行预览）

        Returns:
            工作表概览的文本描述
        """
        summary = f"Sheet: {self.get_name()}\n"
        summary += f"Dimensions: {self.get_row_count()} rows x {self.get_column_count()} columns\n"
        summary += f"Used Range: {self.get_used_range()}\n"

        if include_data:
            summary += "\nData Preview (first 10 rows):\n"
            for i, row in enumerate(self._sheet.iter_rows(max_row=10, values_only=True)):
                if i >= 10:
                    break
                summary += f"Row {i+1}: {row}\n"

        return summary

    def get_cell_range(self, range_str: str) -> str:
        """
        返回范围内单元格的详细信息，包括地址、格式化值、公式（如果有）和样式，
        用于详细检查特定单元格。限制为2000个单元格。

        Args:
            range_str: 单元格范围，如 "A1:C10"；或单个单元格地址，如 "A1"

        Returns:
            单元格详细信息的文本表示
        """
        result = []
        min_col, min_row, max_col, max_row = General._parse_range(range_str)
        # 限制迭代单元格数量以避免过大的输出
        cell_count = (max_row - min_row + 1) * (max_col - min_col + 1)
        if cell_count > 2000:
            return f"Error: The requested range '{range_str}' contains {cell_count} cells, which exceeds the limit of 2000 cells for detailed output. Please specify a smaller range."

        for row_of_cells in self._sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row_of_cells:
                info = f"{cell.coordinate}: "
                info += f"Value={cell.value}, "
                if cell.data_type == 'f':
                    info += f"Formula={cell.value}, "
                info += f"Format={cell.number_format}"
                result.append(info)
        return "\n".join(result)

    def get_raw_cell_data(self, address: str, calc_formula: bool = False) -> Any:
        """
        获取单元格的原始数据值。输出将保持单元格的数据类型（数字、字符串、日期等）。
        注意：openpyxl不会自动将公式转换为值，如果需要设置formula=False读取公式计算结果，必须保证该workbook当前的所有改动均已经save，且执行过calculate方法

        Args:
            address: 要获取数据的地址
            calc_formula: 是否需要返回公式计算后的值

        Returns:
            单元格的原始数据

        Examples:
            >>> sheet.get_raw_cell_data("A1", calc_formula=False)  # 输出: "=SUM(B1:B10)"
            >>> sheet.get_raw_cell_data("A1", calc_formula=True)  # 输出: 100; 前提是该workbook当前的所有改动均已经save，且执行过calculate方法
        """
        cell = self._sheet[address]
        if cell.data_type == 'f' and calc_formula:
            # 如果是公式，且要求返回计算值，则通过data_only=True加载读取
            temp_workbook = load_workbook(self._workbook._filepath, data_only=True)
            temp_sheet = temp_workbook[self._sheet.title]
            temp_value = temp_sheet[address].value
            temp_workbook.close()
            return temp_value
        else:
            # 否则，返回原始值（数值、字符串或公式字符串）
            return cell.value

    def get_raw_range_data(self, range_str: str, calc_formula: bool = False) -> List[List[Any]]:
        """
        与 get_raw_cell_data 相同，但用于单元格范围。输出是数据的二维数组。
        注意：
        1. 即使范围是单个单元格、单行或单列，输出始终是二维数组。
        2. openpyxl不会自动将公式转换为值，如果需要设置formula=False读取公式计算结果，必须保证该workbook当前的所有改动均已经save，且执行过calculate方法


        Args:
            range_str: 单元格范围，如 "A1:B2"，或者单个单元格 "A1"
            calc_formula: 是否需要返回公式计算后的值

        Returns:
            原始数据的二维数组

        Examples:
            >>> sheet.get_raw_range_data("A1:B2", calc_formula=True)
            [[100, 200], [300, 400]]
            >>> sheet.get_raw_range_data("A1:B2", calc_formula=False)
            [["", ""], ["=SUM(B1:B2)", "=SUM(C1:C2)"]]
        """
        need_formula = False
        min_col, min_row, max_col, max_row = General._parse_range(range_str)

        result = []
        for row_of_cells in self._sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            row_data = []
            for original_cell in row_of_cells:
                if calc_formula and original_cell.data_type == 'f':
                    need_formula = True
                    break
                else:
                    row_data.append(original_cell.value)
            if need_formula:
                break
            result.append(row_data)

        if need_formula:
            result = []
            temp_workbook = load_workbook(self._workbook._filepath, data_only=True)
            try:
                temp_calculated_sheet = temp_workbook[self._sheet.title]
                # 遍历原始工作表的范围以获取单元格坐标，然后从 data_only=True 的临时工作表中获取其计算值。
                for row_of_cells in self._sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    row_data = []
                    for original_cell in row_of_cells:
                        # 从 `temp_calculated_sheet` 中使用坐标获取计算值。
                        calculated_cell = temp_calculated_sheet[original_cell.coordinate]
                        row_data.append(calculated_cell.value)
                    result.append(row_data)
            finally:
                # 确保临时工作簿被关闭以释放资源。
                temp_workbook.close()

        return result

    def set_cell(self, address: str, value: Any,
                 number_format: Optional[str] = None):
        """
        设置单元格的值与数字格式，接受 Excel 地址如 "A1"。
        如果是公式，请写成 =SUM(A1:A5) 的格式。

        - 如果需要引用另一个工作表的公式，可以使用格式 "Sheet1!A1"
        - 使用 set_cell 时，请保留数据类型。数字不应加引号：
          正确: sheet.set_cell("A1", 8000)
          错误: sheet.set_cell("A1", "8000")
          公式: sheet.set_cell("A1", "=SUM(B1:B10)")
          数字格式: sheet.set_cell("A1", 8000, number_format=numbers.FORMAT_NUMBER)

        Args:
            address: 单元格地址，如 "A1"
            value: 单元格值（数字、字符串、公式等）
            number_format: 数字格式字符串（例如 numbers.FORMAT_NUMBER、"#,##0.00" 等）

        Example:
            >>> sheet.set_cell("A1", 8000, number_format=numbers.FORMAT_CURRENCY_USD_SIMPLE)
        """
        cell = self._sheet[address]
        cell.value = value

        # 应用数字格式
        if number_format:
            cell.number_format = number_format



    def clear_data(self, range_str: str):
        """
        清除数据但保留样式和格式

        Args:
            range_str: 单元格范围，如 "A1:C10"，或单个单元格地址，如 "A1"
        """
        min_col, min_row, max_col, max_row = General._parse_range(range_str)

        for row_of_cells in self._sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row_of_cells:
                cell.value = None

    def clear_formatting(self, range_str: str):
        """
        清除样式和格式但保留数据

        Args:
            range_str: 单元格范围，如 "A1:C10"，或单个单元格地址，如 "A1"
        """
        min_col, min_row, max_col, max_row = General._parse_range(range_str)

        for row_of_cells in self._sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row_of_cells:
                cell.font = Font()
                cell.border = Border()
                cell.fill = PatternFill()
                cell.number_format = numbers.FORMAT_GENERAL
                cell.alignment = Alignment()

    def get_dependent_cells(self, address: str) -> List[Dict[str, Any]]:
        """
        获取依赖于指定单元格的所有单元格，如公式和引用依赖

        Args:
            address: 单元格地址，如 "A1"

        Returns:
            依赖单元格列表，每个包含 'address' 和 'value' 键
        """
        dependents = []

        for row in self._sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and address in str(cell.value):
                    dependents.append({
                        'address': cell.coordinate,
                        'value': cell.value
                    })

        return dependents

    def get_cell_format(self, address: str) -> str:
        """
        获取单元格的确切格式

        Args:
            address: 单元格地址，如 "A1"

        Returns:
            格式字符串
        """
        cell = self._sheet[address]
        return cell.number_format

    def get_all_specialized_cells(self, address: str) -> str:
        """
        获取按硬编码值、公式、对其他工作表的引用和百分比格式单元格分类的
        单元格的摘要，按行细分。
        用于校对硬编码值，也用于了解要应用格式的范围。

        Args:
            address: 起始地址（通常使用已使用的范围）

        Returns:
            分类摘要的字符串表示
        """
        result = {
            'hardcoded': [],
            'formulas': [],
            'cross_sheet_refs': [],
            'percentages': []
        }

        for row in self._sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    result['formulas'].append(cell.coordinate)
                    if '!' in str(cell.value):
                        result['cross_sheet_refs'].append(cell.coordinate)
                elif cell.value is not None and isinstance(cell.value, (int, float)):
                    result['hardcoded'].append(cell.coordinate)

                if '%' in cell.number_format:
                    result['percentages'].append(cell.coordinate)

        return str(result)

    def set_range_formatting(self, range_str: str, number_format: str):
        """
        批量设置范围的数字格式。
        在设置值时同时设置格式通常更可靠，但对于批量格式化可使用此方法。

        Args:
            range_str: 单元格范围，如 "A1:B10"，或单个单元格地址，如 "A1"
            number_format: 数字格式字符串（例如 numbers.FORMAT_NUMBER、"#,##0.00" 等）

        Example:
            >>> sheet.set_range_formatting("A1:B10", numbers.FORMAT_NUMBER)
        """
        min_col, min_row, max_col, max_row = General._parse_range(range_str)

        for row_of_cells in self._sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row_of_cells:
                cell.number_format = number_format

    def set_style_at(self, range_str: str, style: Dict[str, Any]):
        """
        为指定范围内的单元格应用模块化、非破坏性的样式。
        此方法仅修改您在 style 字典中指定的属性，而不会影响单元格上任何其他已存在的样式。
        例如，设置背景色不会改变现有的字体或边框。

        Args:
            range_str: 单元格范围，如 "A1:B10"，或单个单元格，如 "A1"。
            style: 包含样式属性的字典。支持的键包括：
                   - 'backColor': 背景色，十六进制字符串，如 '#FF0000'。
                   - 'foreColor': 前景色（字体颜色），十六进制字符串，如 '#FFFFFF'。
                   - 'hAlign': 水平对齐 (使用 HorizontalAlign 枚举)。
                   - 'vAlign': 垂直对齐 (使用 VerticalAlign 枚举)。
                   - 'borderTop', 'borderBottom', 'borderLeft', 'borderRight':
                     边框样式。每个键的值都是一个字典，可包含:
                     - 'style': 线条样式 (使用 LineStyle 枚举)。
                     - 'color': 边框颜色 (十六进制字符串)。
                     你可以只提供 'style' 或 'color' 来修改部分边框属性。

        Example:
            # 将 A1:F1 的背景色设为浅灰色，文本居中。不影响单元格已有的字体和边框。
            sheet.set_style_at("A1:F1", {
                'backColor': '#DDDDDD',
                'hAlign': HorizontalAlign.CENTER,
                'vAlign': VerticalAlign.CENTER
            })

            # 为 B5 已有的边框添加一个粗的红色下边框。B5 的其他边框（上、左、右）保持不变。
            sheet.set_style_at("B5", {
                'borderBottom': {'style': LineStyle.THICK, 'color': 'FF0000'}
            })

            # 只改变 C3 单元格的字体颜色为蓝色，不影响背景、对齐等。
            sheet.set_style_at("C3", {'foreColor': '#0000FF'})
        """
        cells_to_process = []
        if ":" in range_str:  # Range like "A1:B10"
            for row_of_cells in self._sheet[range_str]:
                cells_to_process.extend(row_of_cells)
        else:  # Single cell address like "A1"
            cells_to_process.append(self._sheet[range_str])

        for cell in cells_to_process:
            # for cell in row if isinstance(row, tuple) else [row]:
            # 1. 设置字体颜色 (ForeColor)
            if 'foreColor' in style:
                # 总是从现有字体复制，以保留粗体、斜体等其他属性
                font = copy(cell.font)
                font.color = style['foreColor'].replace('#', '')
                cell.font = font

            # 2. 设置背景色 (BackColor)
            if 'backColor' in style:
                # 同样，从现有填充复制，以防有其他填充设置
                fill = copy(cell.fill)
                fill.start_color = style['backColor'].replace('#', '')
                fill.end_color = style['backColor'].replace('#', '')
                fill.fill_type = 'solid'
                cell.fill = fill

            # 3. 设置对齐 (Alignment) - 高效合并处理
            if 'hAlign' in style or 'vAlign' in style:
                alignment = copy(cell.alignment)
                if 'hAlign' in style:
                    if isinstance(style['hAlign'], HorizontalAlign):
                        alignment.horizontal = style['hAlign'].value
                    elif isinstance(style['hAlign'], str):
                        alignment.horizontal = style['hAlign']
                if 'vAlign' in style:
                    if isinstance(style['vAlign'], VerticalAlign):
                        alignment.vertical = style['vAlign'].value
                    elif isinstance(style['vAlign'], str):
                        alignment.vertical = style['vAlign']
                cell.alignment = alignment

            # 4. 设置边框 (Border) - 非破坏性更新
            border_keys = {'borderTop': 'top', 'borderBottom': 'bottom', 'borderLeft': 'left',
                           'borderRight': 'right'}
            # 检查是否有任何边框键在 style 字典中
            if any(key in style for key in border_keys):
                border = copy(cell.border)

                for style_key, border_attr_name in border_keys.items():
                    if style_key in style:
                        border_style_info = style[style_key]  # e.g., {'style': LineStyle.THIN, 'color': 'FF0000'}

                        # 获取现有的 Side 对象，如果不存在则创建一个新的
                        # 这样我们就可以只修改 style 或 color，而不是整个 Side
                        side = getattr(border, border_attr_name)
                        if side is None:
                            side = Side()

                        # 更新线条样式
                        if 'style' in border_style_info:
                            if isinstance(border_style_info['style'], LineStyle):
                                side.style = border_style_info['style'].value
                            elif isinstance(border_style_info['style'], str):
                                side.style = border_style_info['style']

                        # 更新颜色
                        if 'color' in border_style_info:
                            side.color = border_style_info['color'].replace('#', '')

                        # 将更新后的 Side 对象设置回 Border 对象
                        setattr(border, border_attr_name, side)

                cell.border = border

    def set_ib_text_colors(self, range_str: str,
                          ignored_constants: Optional[List[float]] = None,
                          colors: Optional[Dict[str, str]] = None):
        """
        应用投资银行模型文本颜色规范。
        根据单元格内容自动设置文本颜色：
        - 硬编码值（Input）：蓝色 0000FF
        - 公式（Formula）：黑色 000000
        - 跨表引用公式（Cross-Sheet Reference）：绿色 008000

        Args:
            range_str: 单元格范围，如 "A1:B10"，或单个单元格地址，如 "A1"
            ignored_constants: 一个浮点数列表，如果单元格的值等于其中任何一个，
                               则不改变其文本颜色。用于忽略例如 0 的硬编码值。
            colors: 覆盖默认颜色。字典键包括 'inputColor', 'formulaColor', 'crossSheetColor'。
                    值应为 RGB 十六进制字符串，不带 '#'。

        Example:
            >>> # 将 A1:D10 应用投资银行颜色规则，忽略所有值为 0 的硬编码单元格
            >>> sheet.set_ib_text_colors("A1:D10", ignored_constants=[0.0])
            >>> # 应用自定义颜色
            >>> sheet.set_ib_text_colors("A1:D10", colors={
            ...     'inputColor': 'FF0000',  # 红色输入
            ...     'crossSheetColor': '00FF00' # 亮绿色跨表引用
            ... })
        """
        default_colors = {
            'inputColor': '0000FF',  # 蓝色
            'formulaColor': '000000',  # 黑色
            'crossSheetColor': '008000'  # 绿色
        }

        if colors:
            default_colors.update(colors)

        ignored = set(ignored_constants or [])

        min_col, min_row, max_col, max_row = General._parse_range(range_str)

        for row_of_cells in self._sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row_of_cells:
                if isinstance(cell.value, (int, float)) and cell.value not in ignored:
                    font = cell.font.copy() if cell.font else Font()

                    if cell.data_type == 'f':
                        # 公式
                        if '!' in str(cell.value):
                            # 跨表引用
                            font.color = default_colors['crossSheetColor']
                        else:
                            font.color = default_colors['formulaColor']
                    else:
                        # 硬编码值
                        font.color = default_colors['inputColor']

                    cell.font = font

    def set_column_width_at(self, column_letter: str, width: float):
        """
        设置列宽。

        Args:
            column_letter: 列字母，如 "A"
            width: 列宽，以像素为单位。

        Example:
            >>> sheet.set_column_width_at("A", 100.0) # 设置 A 列宽为 100 像素
        """
        # Excel 列宽单位转换：1 单位 ≈ 7 像素
        excel_width = width / 7
        self._sheet.column_dimensions[column_letter].width = excel_width

    def get_column_width_at(self, column_letter: str) -> float:
        """
        获取列宽。

        Args:
            column_letter: 列字母，如 "A"

        Returns:
            列宽，以像素为单位。如果未设置，则返回默认宽度（约 64 像素）。

        Example:
            >>> width_a = sheet.get_column_width_at("A")
            >>> print(f"A 列宽为 {width_a} 像素")
        """
        excel_width = self._sheet.column_dimensions[column_letter].width
        return excel_width * 7 if excel_width else 64  # 默认宽度

    def set_row_height_at(self, row: int, height: float):
        """
        设置行高。

        Args:
            row: 行号（从 1 开始）。
            height: 行高，以像素为单位。

        Example:
            >>> sheet.set_row_height_at(5, 30.0) # 设置第 5 行高为 30 像素
        """
        # Excel 行高单位是点（point），1 点 ≈ 1.33 像素
        point_height = height / 1.33
        self._sheet.row_dimensions[row].height = point_height

    def get_row_height_at(self, row: int) -> float:
        """
        获取行高。

        Args:
            row: 行号（从 1 开始）。

        Returns:
            行高，以像素为单位。如果未设置，则返回默认高度（约 20 像素）。

        Example:
            >>> height_5 = sheet.get_row_height_at(5)
            >>> print(f"第 5 行高为 {height_5} 像素")
        """
        point_height = self._sheet.row_dimensions[row].height
        return point_height * 1.33 if point_height else 15  # 默认高度

class Workbook:
    """工作簿封装类"""

    def __init__(self, filepath: Optional[str] = None):
        """
        创建一个新的工作簿或从现有文件加载工作簿。

        Args:
            filepath: 要加载的 Excel 文件路径。如果为 None，则创建一个新的空工作簿。
        """
        if filepath:
            self._wb = load_workbook(filepath)
        else:
            self._wb = OpenpyxlWorkbook()

        self._filepath = filepath

    def add_sheet(self, title: str, index: Optional[int] = None) -> Worksheet:
        """
        在指定索引位置添加一个新的工作表。基于 openpyxl 的 create_sheet() 方法实现。

        Args:
            title: 新工作表的名称。
            index: 新工作表插入的索引位置（从 0 开始）。如果为 None，添加到末尾。

        Returns:
            新创建的 `Worksheet` 对象。

        Example:
            >>> workbook = Workbook()
            >>> sheet1 = workbook.add_sheet("Sheet1")  # 添加到末尾
            >>> sheet2 = workbook.add_sheet("Sheet2", 0)  # 插入到开头
        """
        # 检查是否已存在同名工作表
        if title in self._wb.sheetnames:
            return self.get_sheet(title)

        ws = self._wb.create_sheet(title=title, index=index)
        return Worksheet(ws, self)

    def remove_sheet(self, sheet: Union[int, str, Worksheet]):
        """
        移除一个工作表。基于 openpyxl 的 remove() 方法实现。

        Args:
            sheet: 可以是：
                - int: 工作表的索引位置（从 0 开始）
                - str: 工作表的名称
                - Worksheet: 工作表对象

        Raises:
            IndexError: 如果索引超出范围。
            KeyError: 如果工作表名称不存在。
            TypeError: 如果参数类型不支持。

        Example:
            >>> workbook.remove_sheet(0)  # 通过索引移除
            >>> workbook.remove_sheet("Sheet1")  # 通过名称移除
            >>> workbook.remove_sheet(sheet_obj)  # 通过对象移除
        """
        if isinstance(sheet, int):
            # 通过索引移除
            if sheet < 0 or sheet >= len(self._wb.worksheets):
                raise IndexError(f"Sheet index {sheet} out of range")
            ws = self._wb.worksheets[sheet]
        elif isinstance(sheet, str):
            # 通过名称移除
            ws = self._wb[sheet]
        elif isinstance(sheet, Worksheet):
            # 通过 Worksheet 对象移除
            ws = sheet._sheet
        else:
            raise TypeError(f"Unsupported type for sheet: {type(sheet)}")

        self._wb.remove(ws)

    def get_sheet(self, sheet: Union[int, str]) -> Worksheet:
        """
        获取一个工作表（支持索引和名称）。基于 openpyxl 的 worksheets 属性实现。

        Args:
            sheet: 可以是：
                - int: 工作表的索引位置（从 0 开始）
                - str: 工作表的名称

        Returns:
            对应的 `Worksheet` 对象。

        Raises:
            IndexError: 如果索引超出范围。
            KeyError: 如果工作表名称不存在。
            TypeError: 如果参数类型不支持。

        Example:
            >>> sheet = workbook.get_sheet(0)  # 通过索引获取
            >>> sheet = workbook.get_sheet("MyData")  # 通过名称获取
        """
        if isinstance(sheet, int):
            # 通过索引获取
            if sheet < 0 or sheet >= len(self._wb.worksheets):
                raise IndexError(f"Sheet index {sheet} out of range")
            return Worksheet(self._wb.worksheets[sheet], self)
        elif isinstance(sheet, str):
            # 通过名称获取
            return Worksheet(self._wb[sheet], self)
        else:
            raise TypeError(f"Unsupported type for sheet: {type(sheet)}")

    def move_sheet(self, sheet: Union[int, str, Worksheet], offset: int):
        """
        移动工作表的位置。基于 openpyxl 的 move_sheet() 方法实现。

        Args:
            sheet: 要移动的工作表，可以是：
                - int: 工作表的索引位置（从 0 开始）
                - str: 工作表的名称
                - Worksheet: 工作表对象
            offset: 相对于当前位置的偏移量（正数向后移动，负数向前移动）。

        Raises:
            IndexError: 如果索引超出范围。
            KeyError: 如果工作表名称不存在。
            TypeError: 如果参数类型不支持。

        Example:
            >>> workbook.move_sheet(0, 2)  # 通过索引，向后移动 2 个位置
            >>> workbook.move_sheet("Sheet1", -1)  # 通过名称，向前移动 1 个位置
            >>> workbook.move_sheet(sheet_obj, 1)  # 通过对象移动
        """
        if isinstance(sheet, int):
            # 通过索引移动
            if sheet < 0 or sheet >= len(self._wb.worksheets):
                raise IndexError(f"Sheet index {sheet} out of range")
            ws = self._wb.worksheets[sheet]
        elif isinstance(sheet, str):
            # 通过名称移动
            ws = self._wb[sheet]
        elif isinstance(sheet, Worksheet):
            # 通过对象移动
            ws = sheet._sheet
        else:
            raise TypeError(f"Unsupported type for sheet: {type(sheet)}")

        self._wb.move_sheet(ws, offset=offset)

    def copy_sheet(self, sheet: Union[int, str, Worksheet]) -> Worksheet:
        """
        复制一个工作表。基于 openpyxl 的 copy_worksheet() 方法实现。

        Warning:
            此功能只能在同一工作簿内复制工作表，不能跨工作簿复制。

        Args:
            sheet: 要复制的工作表，可以是：
                - int: 工作表的索引位置（从 0 开始）
                - str: 工作表的名称
                - Worksheet: 工作表对象

        Returns:
            复制后的新 `Worksheet` 对象。

        Raises:
            IndexError: 如果索引超出范围。
            KeyError: 如果工作表名称不存在。
            TypeError: 如果参数类型不支持。

        Example:
            >>> original = workbook.get_sheet(0)
            >>> copied = workbook.copy_sheet(original)  # 通过对象复制
            >>> copied2 = workbook.copy_sheet("Sheet1")  # 通过名称复制
            >>> copied3 = workbook.copy_sheet(0)  # 通过索引复制
        """
        if isinstance(sheet, int):
            # 通过索引复制
            if sheet < 0 or sheet >= len(self._wb.worksheets):
                raise IndexError(f"Sheet index {sheet} out of range")
            ws = self._wb.worksheets[sheet]
        elif isinstance(sheet, str):
            # 通过名称复制
            ws = self._wb[sheet]
        elif isinstance(sheet, Worksheet):
            # 通过对象复制
            ws = sheet._sheet
        else:
            raise TypeError(f"Unsupported type for sheet: {type(sheet)}")

        copied_ws = self._wb.copy_worksheet(ws)
        return Worksheet(copied_ws, self)

    def get_sheet_count(self) -> int:
        """
        获取工作簿中工作表的总数。基于 openpyxl 的 worksheets 属性实现。

        Returns:
            工作表的数量。

        Example:
            >>> count = workbook.get_sheet_count()
            >>> print(f"工作簿中有 {count} 个工作表")
        """
        return len(self._wb.worksheets)

    def get_sheet_names(self) -> list:
        """
        获取工作簿中所有工作表的名称列表。基于 openpyxl 的 sheetnames 属性实现。

        Returns:
            工作表名称的列表（按顺序）。

        Example:
            >>> names = workbook.get_sheet_names()
            >>> print(names)  # ['Sheet1', 'Sheet2', 'MyData']
        """
        return self._wb.sheetnames

    def set_active_sheet(self, sheet: Union[int, str, Worksheet]):
        """
        设置当前活动（可见）的工作表。基于 openpyxl 的 active 属性实现。

        Args:
            sheet: 要设置为活动的工作表，可以是：
                - int: 工作表的索引位置（从 0 开始）
                - str: 工作表的名称
                - Worksheet: 工作表对象

        Raises:
            IndexError: 如果索引超出范围。
            KeyError: 如果工作表名称不存在。
            TypeError: 如果参数类型不支持。

        Example:
            >>> workbook.set_active_sheet(0)  # 通过索引设置
            >>> workbook.set_active_sheet("Sheet1")  # 通过名称设置
            >>> workbook.set_active_sheet(sheet_obj)  # 通过对象设置
        """
        if isinstance(sheet, int):
            # 通过索引设置
            if sheet < 0 or sheet >= len(self._wb.worksheets):
                raise IndexError(f"Sheet index {sheet} out of range")
            # openpyxl 的 active 可以直接接受索引
            self._wb.active = sheet
        elif isinstance(sheet, str):
            # 通过名称设置，获取对应的工作表对象
            ws = self._wb[sheet]
            # openpyxl 的 active 可以接受 Worksheet 对象
            self._wb.active = ws
        elif isinstance(sheet, Worksheet):
            # 通过对象设置
            # openpyxl 的 active 可以接受 Worksheet 对象
            self._wb.active = sheet._sheet
        else:
            raise TypeError(f"Unsupported type for sheet: {type(sheet)}")

    def get_active_sheet(self) -> Worksheet:
        """
        获取当前活动的工作表。基于 openpyxl 的 active 属性实现。

        Returns:
            当前活动的 `Worksheet` 对象。

        Example:
            >>> active_sheet = workbook.get_active_sheet()
            >>> print(f"活动工作表是: {active_sheet.get_name()}")
        """
        return Worksheet(self._wb.active, self)

    def set_gridlines_for_all_sheets(self, show_gridlines: bool):
        """
        设置工作簿中所有工作表的网格线是否可见。

        Args:
            show_gridlines: 如果为 True，则显示网格线；如果为 False，则隐藏网格线。

        Example:
            >>> workbook.set_gridlines_for_all_sheets(False)  # 隐藏所有工作表的网格线
        """
        for sheet in self._wb.worksheets:
            sheet.sheet_view.showGridLines = show_gridlines

    def auto_resize_columns(self):
        """
        自动调整工作簿中所有工作表的列宽，以适应其内容。

        此方法会遍历工作簿中的所有工作表，并为每个工作表的每一列计算最佳宽度。
        它会根据单元格内容的长度、字体大小和加粗情况来计算列宽。
        """
        for sheet_name in self._wb.sheetnames:
            worksheet = self._wb[sheet_name]
            column_widths = {}
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        column_letter = get_column_letter(cell.column)
                        # 格式化显示值
                        cell_text = format_cell_value(cell)
                        # 获取字体属性
                        font_size = float(cell.font.size) if cell.font and cell.font.size else 11.0
                        bold = bool(cell.font.bold) if cell.font else False
                        cell_length = calculate_display_width(cell_text, font_size, bold)
                        if column_letter not in column_widths:
                            column_widths[column_letter] = cell_length
                        else:
                            column_widths[column_letter] = max(
                                column_widths[column_letter],
                                cell_length
                            )

            # 应用列宽
            for column_letter, width in column_widths.items():
                adjusted_width = width + 2  # buffer ≈ 20 像素
                adjusted_width = max(adjusted_width + 2, 8)  # padding
                adjusted_width = min(adjusted_width, 50)  # 限制最大宽度
                adjusted_width = round(adjusted_width, 1)

                worksheet.column_dimensions[column_letter].width = adjusted_width
                print(f"Sheet {sheet_name} Column {column_letter}: width set to {adjusted_width}")

        print("  Auto resize column width successfully!")

    def calculate(self) -> str:
        """
        重新计算工作簿中所有公式的值。此功能依赖的 'recalc' 库已安装。

        Returns:
            一个包含计算结果和任何发现的 Excel 错误的字典序列化后的str。
        示例成功返回:
        {
            "status": "success",
            "total_errors": 0,
            "error_summary": {},
            "total_formulas": 123
        }
        示例包含错误返回:
        {
            "status": "errors_found",
            "total_errors": 2,
            "error_summary": {
                "#DIV/0!": {
                    "count": 1,
                    "locations": ["Sheet1!A1"]
                },
                "#VALUE!": {
                    "count": 1,
                    "locations": ["Sheet2!C5"]
                }
            },
            "total_formulas": 123
        }
        示例失败返回 (e.g., 文件不存在或LibreOffice宏配置失败):
        {
            "error": "File /path/to/nonexistent.xlsx does not exist"
        }
        """
        self.save()
        first_sheet_original_name = None
        if len(self._wb.worksheets) > 0:
            first_sheet_original_name = self._wb.worksheets[0].title
        result = recalc(self._filepath) # recalc 这里可能会因为LibreOffice的Bug把第一个sheet1改成sheet1 -1
        self._wb = load_workbook(self._filepath)
        if first_sheet_original_name == "sheet1" and len(self._wb.worksheets) > 0:
            current_first_sheet_name = self._wb.worksheets[0].title
            if current_first_sheet_name != first_sheet_original_name:
                self._wb.worksheets[0].title = first_sheet_original_name
                self.save()
        return json.dumps(result, indent=2)

    def save(self, filepath: Optional[str] = None):
        """
        保存工作簿到指定文件路径。如果未提供路径，则保存到初始化时提供的路径。

        Args:
            filepath: 保存工作簿的文件路径。如果为 None，则使用初始化时提供的路径。

        Raises:
            ValueError: 如果未提供文件路径且工作簿未曾保存过。

        Example:
            >>> workbook.save("my_new_workbook.xlsx")
            >>> workbook.save() # 如果已通过文件路径初始化工作簿
        """
        save_path = filepath or self._filepath
        if not save_path:
            raise ValueError("未指定保存路径")
        self._wb.save(save_path)

    def close(self):
        """
        关闭工作簿。
        此方法释放与工作簿相关的任何资源。建议在完成工作后调用此方法。

        Example:
            >>> workbook.close()
        """
        self._wb.close()


class General:
    """通用工具类"""

    @staticmethod
    def copy_file(file_path: str, new_file_path: str):
        """
        将指定的文件（例如Excel文件）复制到新位置。
        此操作包括复制文件数据和元数据（例如创建/修改时间）。

        Args:
            file_path: 要复制的源文件的完整路径。
            new_file_path: 新文件的目标路径和文件名。

        Raises:
            FileNotFoundError: 如果源文件不存在。
            IOError: 如果复制操作失败（例如权限问题）。

        Example:
            >>> General.copy_file("/app/data/files/original.xlsx", "/app/data/files/copy_of_original.xlsx")
        """
        try:
            shutil.copy2(file_path, new_file_path)
            print(f"File '{file_path}' successfully copied to '{new_file_path}'")
        except FileNotFoundError:
            raise FileNotFoundError(f"Source file not found: {file_path}")
        except Exception as e:
            raise IOError(f"Failed to copy file '{file_path}' to '{new_file_path}': {e}")

    @staticmethod
    def copy_range_from_workbook(from_workbook: Workbook, from_range: str,
                                 to_workbook: Workbook, to_range: str):
        """
        将一个工作簿中的指定单元格范围（包括值、公式、格式、样式、列宽、行高和合并单元格）
        完整复制到另一个工作簿的指定目标范围。

        Args:
            from_workbook: 源 `Workbook` 对象。
            from_range: 源范围字符串，必须包含工作表名称，例如 "Sheet1!A1:D10"。
            to_workbook: 目标 `Workbook` 对象。
            to_range: 目标范围字符串，必须包含工作表名称，例如 "Sheet2!B2"。
                      目标范围的起始单元格决定了内容的放置位置。注意，和from_range参数不同，这里只需要左上角单元格的位置

        Raises:
            ValueError: 如果源或目标范围字符串格式不正确，未包含工作表名称。

        Example:
            >>> src_wb = Workbook("source.xlsx")
            >>> dest_wb = Workbook()
            >>> General.copy_range_from_workbook(src_wb, "DataSheet!A1:C5", dest_wb, "Report!E10")
            >>> dest_wb.save("destination.xlsx")
        """
        # 解析源和目标范围（包含工作表名）
        if '!' not in from_range or '!' not in to_range:
            raise ValueError("范围必须包含工作表名称，格式: 'Sheet1!A1:D10'")

        from_sheet_name, from_cells = from_range.split('!')
        to_sheet_name, to_cells = to_range.split('!')

        from_sheet = from_workbook.get_sheet(from_sheet_name)
        to_sheet = to_workbook.get_sheet(to_sheet_name)

        # 解析单元格范围
        from_min_col, from_min_row, from_max_col, from_max_row = General._parse_range(from_cells)
        to_start_col, to_start_row = General._parse_cell(to_cells.split(':')[0])

        # 计算目标范围
        rows_count = from_max_row - from_min_row + 1
        cols_count = from_max_col - from_min_col + 1
        to_max_row = to_start_row + rows_count - 1
        to_max_col = to_start_col + cols_count - 1

        # 1. 复制单元格数据、格式和样式
        for row_offset in range(rows_count):
            for col_offset in range(cols_count):
                from_row = from_min_row + row_offset
                from_col = from_min_col + col_offset
                to_row = to_start_row + row_offset
                to_col = to_start_col + col_offset

                # 获取源单元格和目标单元格
                from_cell = from_sheet._sheet.cell(row=from_row, column=from_col)
                to_cell = to_sheet._sheet.cell(row=to_row, column=to_col)

                # 复制值
                to_cell.value = from_cell.value

                # 完整复制样式
                if from_cell.has_style:
                    # 字体
                    to_cell.font = copy(from_cell.font)
                    # 填充（背景色）
                    to_cell.fill = copy(from_cell.fill)
                    # 边框
                    to_cell.border = copy(from_cell.border)
                    # 对齐
                    to_cell.alignment = copy(from_cell.alignment)
                    # 数字格式
                    to_cell.number_format = copy(from_cell.number_format)
                    # 保护
                    to_cell.protection = copy(from_cell.protection)
                    # to_cell._style = copy(from_cell._style)

        # 2. 复制列宽
        for col_offset in range(cols_count):
            from_col = from_min_col + col_offset
            to_col = to_start_col + col_offset

            from_col_letter = get_column_letter(from_col)
            to_col_letter = get_column_letter(to_col)

            # 获取源列宽
            from_col_dim = from_sheet._sheet.column_dimensions.get(from_col_letter)
            if from_col_dim and from_col_dim.width:
                # 复制列宽
                to_sheet._sheet.column_dimensions[to_col_letter].width = from_col_dim.width

        # 3. 复制行高
        for row_offset in range(rows_count):
            from_row = from_min_row + row_offset
            to_row = to_start_row + row_offset

            # 获取源行高
            from_row_dim = from_sheet._sheet.row_dimensions.get(from_row)
            if from_row_dim and from_row_dim.height:
                # 复制行高
                to_sheet._sheet.row_dimensions[to_row].height = from_row_dim.height

        # 4. 复制合并单元格
        # 遍历源范围内的所有合并单元格
        for merged_range in from_sheet._sheet.merged_cells.ranges:
            # 检查合并单元格是否在源范围内
            if General._is_range_in_range(merged_range, from_min_col, from_min_row, from_max_col, from_max_row):
                # 计算在目标工作表中的位置
                merge_min_col = merged_range.min_col - from_min_col + to_start_col
                merge_min_row = merged_range.min_row - from_min_row + to_start_row
                merge_max_col = merged_range.max_col - from_min_col + to_start_col
                merge_max_row = merged_range.max_row - from_min_row + to_start_row

                # 在目标工作表中合并对应的单元格
                to_merge_range = f"{get_column_letter(merge_min_col)}{merge_min_row}:{get_column_letter(merge_max_col)}{merge_max_row}"
                to_sheet._sheet.merge_cells(to_merge_range)

    @staticmethod
    def _parse_range(range_str: str) -> Tuple[int, int, int, int]:
        """解析范围字符串，返回 (min_col, min_row, max_col, max_row)"""
        if ':' in range_str:
            start, end = range_str.split(':')
            start_col, start_row = General._parse_cell(start)
            end_col, end_row = General._parse_cell(end)
            return start_col, start_row, end_col, end_row
        else:
            col, row = General._parse_cell(range_str)
            return col, row, col, row

    @staticmethod
    def _parse_cell(cell_str: str) -> Tuple[int, int]:
        """解析单元格地址，返回 (col, row)"""
        col_str, row_str = coordinate_from_string(cell_str)
        col = column_index_from_string(col_str)
        row = int(row_str)
        return col, row

    @staticmethod
    def _is_range_in_range(merged_range, min_col: int, min_row: int, max_col: int, max_row: int) -> bool:
        """检查合并单元格范围是否在指定范围内"""
        return (merged_range.min_col >= min_col and
                merged_range.max_col <= max_col and
                merged_range.min_row >= min_row and
                merged_range.max_row <= max_row)

    @staticmethod
    def get_api_info(function_name: str) -> str:
        """
        获取指定 API 函数的文档字符串。

        Args:
            function_name: 要获取文档的函数名称（字符串）。
                            例如 "Workbook.add_sheet" 或 "Worksheet.set_cell"。

        Returns:
            函数的文档字符串。如果找不到函数，则返回 "未找到该函数的文档"。

        Example:
            >>> print(General.get_api_info("Workbook.add_sheet"))
        """
        # 将函数名分割为类名和方法名
        parts = function_name.split('.')
        if len(parts) == 2:
            class_name, method_name = parts[0], parts[1]
            target_class = None
            if class_name == "Workbook":
                target_class = Workbook
            elif class_name == "Worksheet":
                target_class = Worksheet
            elif class_name == "General":
                target_class = General
            # 可以根据需要添加其他类

            if target_class:
                method = getattr(target_class, method_name, None)
                if method and method.__doc__:
                    return method.__doc__.strip()
        elif len(parts) == 1: # 可能是顶级函数或类本身
            # 检查是否是类本身的文档
            if function_name == "Workbook":
                return Workbook.__doc__.strip() if Workbook.__doc__ else "无文档"
            elif function_name == "Worksheet":
                return Worksheet.__doc__.strip() if Worksheet.__doc__ else "无文档"
            elif function_name == "General":
                return General.__doc__.strip() if General.__doc__ else "无文档"

        return "未找到该函数的文档"


# 全局断言函数
def assert_condition(condition: bool, message: str):
    """断言条件，失败时抛出异常"""
    if not condition:
        raise AssertionError(message)
